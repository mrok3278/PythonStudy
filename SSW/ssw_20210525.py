'''
엑셀 구조
    Application > Workbook(File) > Sheets > Rows, Columns, Range, Cells, ...
'''

from openpyxl import Workbook, load_workbook   # pip install openpyxl
import datetime

# 데이터 쓰기
wb = Workbook() # 워크북 생성
ws = wb.active  # 활성화된 시트 선택
ws.title = "sample"     # 시트명 지정

# 셀 주소로 값 입력
ws['A1'] = 'SSW says'

# 셀 좌표로 값 입력(행, 열)
ws.cell(1, 2, 'Slow and Steady Wins the race.')

# 현재 시각 입력
ws['A2'] = datetime.datetime.now()

# 행 단위로 값 입력(마지막에 행 추가)
ws.append([1, 2, 3])
ws.append([4, 5, 6])
ws.append([7, 8, 9])

# 수식 입력
ws.cell(3, 4, '=SUM(A3:C3)')
ws.cell(4, 4, '=SUM(A4:C4)')
ws.cell(5, 4, '=SUM(A5:C5)')

# 엑셀 파일 저장
wb.save("sample1.xlsx")


'''
데이터 읽기
'''

wb = load_workbook("sample1.xlsx", data_only=False)  # data_only=True: 수식은 제외하고 값만 불러오기
ws = wb['sample']  # 시트 이름으로 불러오기

# 셀 주소로 값 출력
print(ws['A1'].value)
print()

# 셀 좌표로 값 출력(행, 열)
print(ws.cell(1, 2).value)
print()

# 지정한 셀의 값 출력
cells = ws['A3':'D5']
for row in cells:   # 행
    for cell in row:    # 열
        print(cell.value)
print()

# 모든 행 단위로 출력
for row in ws.rows:
    print(row)
print()

# 모든 열 단위로 출력
for column in ws.columns:
    print(column)
print()

# 모든 행과 열 출력
all_values = []
for row in ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)
