'''
웹 스크래핑과 웹 크롤링

- 웹 스크래핑
    * 웹 사이트 상에서 원하는 정보를 추출하는 기술

- 웹 크롤링
    * 웹 크롤러(자동화 봇)가 일정 규칙으로 웹 페이지를 브라우징하는 것

- 주의점
    * 서버에 과부하를 주는 등의 피해를 줄 수 있으므로, 법적인 문제에 대해 전문적으로 알아볼 것
    * 특히, 학습용이 아닌 상용을 목적으로 크롤링 및 스크래핑시에는 더욱 주의할 것
    * robots.txt 에서 크롤링 가능 여부 확인
        .검색엔진 크롤링 봇들이 해당 파일 조회 후 수집 여부 결정
        .참조: https://searchadvisor.naver.com/guide/seo-basic-robots
'''

import requests    # 웹 페이지 접속
from bs4 import BeautifulSoup   # HTML 문서 파싱(구문 분석)
import openpyxl

wb = openpyxl.Workbook()    # 워크북 생성

res = requests.get('https://unse.daily.co.kr/?p=zodiac')
print(res.status_code)
# print(res.text)

bs = BeautifulSoup(res.text, 'html.parser')    # 구문분석기: 'lxml', 'html.parser', ...

# 부모 요소 찾기
section = bs.select_one('#card')

# 자식 요소 찾기
uls = section.select('ul')
for ul in uls:
    cnt = 0
    lis = ul.select('li')
    for li in lis:
        cnt += 1    # cnt = cnt + 1 과 동일
        
        if cnt == 1:    # 첫번째 li
            name = li.select_one('div.txt_box > b').get_text()
            ws = wb.create_sheet(name)

            summary = li.select_one('div.txt_box > p').get_text()
            print(name, summary)
            ws.append(['종합운세', summary])

            # append 함수 이용하는 방법
            # ws.append([])
            # ws.append(['연도', '운세'])

            # cell 함수 이용하는 방법
            ws.cell(3, 1, '연도')
            ws.cell(3, 2, '운세')
        else:           # 나머지 li
            year = li.select_one('span').get_text()
            fortune = li.select_one('p').get_text()
            print(year, fortune)
            ws.append([year, fortune])
        print()

fortune_date = bs.select_one('#unse_date_menu > div > a.on').get_text()
# 퀴즈: 오늘 (05.25) --> (0525) 출력해보세요.
print(fortune_date)

wb.remove(wb['Sheet'])
wb.save('fortune.xlsx')
