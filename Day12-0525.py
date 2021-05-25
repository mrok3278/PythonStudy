# Day12-0525.py

from openpyxl import Workbook, load_workbook   # pip install openpyxl
import datetime

# 데이터 쓰기
wb = Workbook() # 워크북 생성
ws = wb.active  # 활성화된 시트 선택

# 셀 주소로 값 입력
ws['A1'] = 'Hi'

# 셀 좌표로 값 입력(행, 열)
ws.cell(1, 2, 'Slow and Steady Wins the race.')

# 현재 시각 입력
ws['A2'] = datetime.datetime.now()

# 행 단위로 값 입력(마지막에 행 추가)
ws.append([1, 2, 3])
ws.append([4, 5, 6])
ws.append([7, 8, 9])

# 수식 입력
ws.cell(3, 4, '=SUM(A3:C3)')
ws.cell(4, 4, '=SUM(A4:C4)')
ws.cell(5, 4, '=SUM(A5:C5)')

# 엑셀 파일 저장
wb.save('Resource/sample1.xlsx')
wb.close()

wb = load_workbook('Resource/sample1.xlsx', data_only=True)
ws = wb['Sheet']

for row in ws.rows:
    
    row_value = []
    
    for cell in row:
        
        if cell.value == None:
            continue
                
        row_value.append(cell.value)
        
    print(row_value)
    
    wb.close()
    