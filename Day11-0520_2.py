# Day11-0520.py

import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image

wb = Workbook()
ws = wb.active

res = requests.get('https://unse.daily.co.kr/?p=zodiac')
bs = BeautifulSoup(res.text, 'html.parser')

section = bs.select_one('#card')
uls = section.select('ul')

year = ''
text = ''

#code_1 > li.start_li > div.txt_box > b
for ul in uls:
    
    year = ul.select_one('li.start_li > div.txt_box > b').get_text()
    text = ul.select_one('li.start_li > div.txt_box > p').get_text()
    
    ws = wb.create_sheet(year)
    
    try:
        img = Image(f"Resource/{year}.jpg")
        ws.add_image(img, 'A1')   # B2 셀에 지정한 이미지 삽입
        
    except Exception as e:
        ws['A1'] = '종합운세'

    ws['B1'] = text
    
    # ws.append([''])
    # ws.append(['연도','운세'])
    ws.cell(3,1, '연도')
    ws.cell(3,2, '운세')
              
    lis = ul.select("li")
    
    for li in lis:
        
        year = li.select_one("span").get_text()
        text = li.select_one("p").get_text()
        
        if not bool(year):
            continue
        
        ws.append([year, text])

unse_date = ''
#unse_box > section > section:nth-child(2) > h3
h3s = bs.select('h3')
for h3 in h3s:

    if h3.get('class')[0] != 'tit_date':
        continue
    
    unse_date = h3.get_text()
    print(unse_date)
    break

wb.remove(wb['Sheet'])    
wb.save('Resource/%s.xlsx' %(unse_date))
